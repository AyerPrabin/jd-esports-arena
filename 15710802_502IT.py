"""
JD Systems Toolkit -- Algorithms & Data Structures, applied to real JD Esports
Arena / JD Stock problems
================================================================================
Originally written as a university Algorithms & Data Structures assessment
(five independent systems, each mapped to one problem from the module brief);
repurposed here with real data so it's an actual working toolkit rather than a
textbook demo. Strictly Object-Oriented. Core algorithms (heap, binary search,
nearest neighbour, Jaccard similarity, sorting) are hand-written rather than
imported (no heapq, no bisect). Every structural operation prints a
step-by-step trace so console output can double as an audit log.

Problem 1: JD Stock delivery routes      -> Graph + Nearest Neighbour TSP heuristic
Problem 2: Zulu server job scheduling    -> Max-Heap priority queue + FCFS/SJF/Round Robin
Problem 3: Tournament practice bookings  -> Sorted array + manual Binary Search
Problem 4: JD Stock parts inventory      -> Hash table + sorting/search/auto-reorder
Problem 5: JD Stock parts recommender    -> Sets + Jaccard similarity collaborative filtering

Note: "JD Stock" here is illustrative sample data (Nepal bike/car parts, in the
same spirit as the assessment's original Spark_Plugs/Brake_Pads example) -- it
is not connected to any live system or real supplier catalog.

Run `python 15710802_502IT.py` for the interactive menu.
Run `python -m unittest 15710802_502IT -v` for the correctness test suite.
"""
import os
import unittest

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "15710802_502IT_data.xlsx")


# ==========================================================
# PROBLEM 1: JD STOCK DELIVERY ROUTES (GRAPH + NEAREST NEIGHBOUR)
# ==========================================================
class Location:
    def __init__(self, location_id, name, x, y):
        self.location_id = location_id
        self.name = name
        self.x = float(x)
        self.y = float(y)

    def __str__(self):
        return f"{self.name}({self.x:.0f},{self.y:.0f})"


class DeliveryRouteOptimizer:
    """Locations form a complete graph (every pair connected by a straight-line
    distance). Nearest Neighbour is a greedy O(n^2) heuristic approximation
    of the Travelling Salesperson Problem.
    """

    def __init__(self):
        self.locations = {}  # location_id -> Location

    def add_location(self, location):
        self.locations[location.location_id] = location
        print(f"[GRAPH] Added node '{location.location_id}' -> {location}")

    def _distance(self, a, b):
        return ((a.x - b.x) ** 2 + (a.y - b.y) ** 2) ** 0.5

    def nearest_neighbour_route(self, start_id):
        if start_id not in self.locations:
            print(f"[ERROR] Start location '{start_id}' not found.")
            return None, 0.0

        unvisited = set(self.locations.keys())
        unvisited.remove(start_id)
        route = [start_id]
        current = start_id
        total_distance = 0.0
        comparisons = 0

        print(f"[NN-TSP] Starting route at '{start_id}'.")
        while unvisited:
            nearest_id = None
            nearest_dist = float("inf")
            for candidate_id in unvisited:
                comparisons += 1
                dist = self._distance(self.locations[current], self.locations[candidate_id])
                print(f"   [COMPARE] {current} -> {candidate_id}: distance={dist:.2f}")
                if dist < nearest_dist:
                    nearest_dist = dist
                    nearest_id = candidate_id
            print(f"   [CHOOSE] Nearest unvisited from '{current}' is '{nearest_id}' "
                  f"(distance={nearest_dist:.2f})")
            route.append(nearest_id)
            total_distance += nearest_dist
            unvisited.remove(nearest_id)
            current = nearest_id

        return_leg = self._distance(self.locations[current], self.locations[start_id])
        total_distance += return_leg
        route.append(start_id)
        print(f"   [RETURN] {current} -> {start_id}: distance={return_leg:.2f}")

        n = len(self.locations)
        print(f"[COMPLEXITY] Nearest Neighbour used {comparisons} distance comparisons for "
              f"n={n} nodes (upper bound n*(n-1)/2 = {n * (n - 1) // 2}) -> O(n^2).")
        return route, total_distance

    def brute_force_optimal(self, start_id, node_limit=8):
        """Exact TSP via exhaustive permutation search. O(n!) -- only safe for
        small n. Used to benchmark how close the heuristic gets to optimal.
        """
        import itertools
        n = len(self.locations)
        if n > node_limit:
            print(f"[BRUTE FORCE] Skipped: n={n} exceeds safe limit of {node_limit} (O(n!) blow-up).")
            return None, None

        others = [lid for lid in self.locations if lid != start_id]
        best_route, best_distance, tried = None, float("inf"), 0
        for perm in itertools.permutations(others):
            tried += 1
            route = [start_id] + list(perm) + [start_id]
            dist = sum(self._distance(self.locations[route[i]], self.locations[route[i + 1]])
                       for i in range(len(route) - 1))
            if dist < best_distance:
                best_distance, best_route = dist, route
        print(f"[BRUTE FORCE] Evaluated {tried} permutations ((n-1)!) for the exact optimum -> O(n!).")
        return best_route, best_distance


# ==========================================================
# PROBLEM 2: ZULU SERVER JOB SCHEDULING (MAX-HEAP + FCFS/SJF/ROUND ROBIN)
# ==========================================================
class Process:
    def __init__(self, pid, name, priority, burst_time, cores_required, memory_required):
        self.pid = pid
        self.name = name
        self.priority = int(priority)
        self.burst_time = int(burst_time)          # ms of CPU time needed
        self.cores_required = int(cores_required)
        self.memory_required = int(memory_required)  # MB

    def __str__(self):
        return (f"{self.name}(PID{self.pid}, P{self.priority}, burst={self.burst_time}ms, "
                f"cores={self.cores_required}, mem={self.memory_required}MB)")


class ResourcePool:
    def __init__(self, total_cores, total_memory):
        self.total_cores = total_cores
        self.total_memory = total_memory
        self.available_cores = total_cores
        self.available_memory = total_memory

    def can_allocate(self, process):
        return (process.cores_required <= self.available_cores
                and process.memory_required <= self.available_memory)

    def allocate(self, process):
        self.available_cores -= process.cores_required
        self.available_memory -= process.memory_required

    def release(self, process):
        self.available_cores += process.cores_required
        self.available_memory += process.memory_required

    def __str__(self):
        return f"Cores {self.available_cores}/{self.total_cores}, Memory {self.available_memory}/{self.total_memory}MB"


class PriorityScheduler:
    """Binary max-heap stored in a flat Python list. Works on any object
    exposing a numeric `.priority` attribute -- reused here to order
    incoming Processes by priority before resources are allocated.
    Parent of index i is (i-1)//2. Children of index i are 2i+1, 2i+2.
    """

    def __init__(self):
        self.heap = []

    def insert(self, task):
        self.heap.append(task)
        index = len(self.heap) - 1
        print(f"[INSERT] Placed {task} at index {index} (bottom of heap).")
        self._heapify_up(index)

    def _heapify_up(self, index):
        parent_index = (index - 1) // 2
        while index > 0 and self.heap[parent_index].priority < self.heap[index].priority:
            child = self.heap[index]
            parent = self.heap[parent_index]
            print(f"   [HEAPIFY-UP] {child} at index {index} > parent {parent} "
                  f"at index {parent_index} -> SWAP")
            self.heap[index], self.heap[parent_index] = self.heap[parent_index], self.heap[index]
            index = parent_index
            parent_index = (index - 1) // 2
        if index == 0:
            print(f"   [HEAPIFY-UP] {self.heap[0]} has reached the ROOT (index 0). Bubble-up complete.")
        else:
            print(f"   [HEAPIFY-UP] Heap property satisfied at index {index}. Bubble-up complete.")

    def extract_max(self):
        if len(self.heap) == 0:
            print("[EXTRACT] Heap is empty. Nothing to extract.")
            return None
        if len(self.heap) == 1:
            removed = self.heap.pop()
            print(f"[EXTRACT] Removed last remaining item: {removed}")
            return removed

        max_item = self.heap[0]
        last_item = self.heap.pop()
        print(f"[EXTRACT] Removing root {max_item}. Moving last element {last_item} "
              f"from index {len(self.heap)} to root (index 0).")
        self.heap[0] = last_item
        self._heapify_down(0)
        return max_item

    def _heapify_down(self, index):
        size = len(self.heap)
        while True:
            largest = index
            left = 2 * index + 1
            right = 2 * index + 2

            if left < size and self.heap[left].priority > self.heap[largest].priority:
                largest = left
            if right < size and self.heap[right].priority > self.heap[largest].priority:
                largest = right

            if largest != index:
                print(f"   [HEAPIFY-DOWN] {self.heap[largest]} at index {largest} > "
                      f"{self.heap[index]} at index {index} -> SWAP")
                self.heap[index], self.heap[largest] = self.heap[largest], self.heap[index]
                index = largest
            else:
                print(f"   [HEAPIFY-DOWN] Heap property satisfied at index {index}. Sift-down complete.")
                break

    def print_queue(self):
        print(f"Current queue: [{', '.join(str(t) for t in self.heap)}]")


class TaskManager:
    """Assigns limited CPU/memory resources to incoming processes, ordered
    by priority (max-heap). Processes that arrive when resources are
    unavailable wait until something is released.
    """

    def __init__(self, total_cores, total_memory):
        self.pool = ResourcePool(total_cores, total_memory)
        self.ready_heap = PriorityScheduler()
        self.waiting = []
        self.running = []
        self._next_pid = 1

    def submit_process(self, name, priority, burst_time, cores_required, memory_required):
        process = Process(self._next_pid, name, priority, burst_time, cores_required, memory_required)
        self._next_pid += 1
        print(f"[SUBMIT] {process.name} (PID {process.pid}) queued with priority {process.priority}.")
        self.ready_heap.insert(process)
        return process

    def dispatch_all(self):
        print(f"[POOL] {self.pool}")
        while self.ready_heap.heap:
            process = self.ready_heap.extract_max()
            if self.pool.can_allocate(process):
                self.pool.allocate(process)
                self.running.append(process)
                print(f"[ALLOCATE] {process.name} granted {process.cores_required} core(s), "
                      f"{process.memory_required}MB. Pool now: {self.pool}")
            else:
                self.waiting.append(process)
                print(f"[WAIT] Not enough resources for {process.name} "
                      f"(needs {process.cores_required} cores/{process.memory_required}MB, "
                      f"pool has {self.pool}). Added to waiting list.")

    def release_process(self, pid):
        for process in self.running:
            if process.pid == pid:
                self.pool.release(process)
                self.running.remove(process)
                print(f"[RELEASE] {process.name} finished, resources returned. Pool now: {self.pool}")
                self._retry_waiting()
                return True
        print(f"[ERROR] PID '{pid}' is not currently running.")
        return False

    def _retry_waiting(self):
        still_waiting = []
        for process in self.waiting:
            if self.pool.can_allocate(process):
                self.pool.allocate(process)
                self.running.append(process)
                print(f"[ALLOCATE] (from wait list) {process.name} granted resources. Pool now: {self.pool}")
            else:
                still_waiting.append(process)
        self.waiting = still_waiting

    # ---- comparison scheduling algorithms (ordering only, no resource pool) ----
    @staticmethod
    def fcfs_schedule(processes):
        print("[FCFS] Execution order = arrival order:")
        clock = 0
        for p in processes:
            print(f"   t={clock:4d} -> {p.name} runs for {p.burst_time}ms")
            clock += p.burst_time
        print(f"[FCFS] Total turnaround time: {clock}ms")
        return clock

    @staticmethod
    def sjf_schedule(processes):
        """Non-preemptive Shortest-Job-First, ordered via manual selection sort."""
        remaining = list(processes)
        order = []
        while remaining:
            shortest = remaining[0]
            for p in remaining:
                if p.burst_time < shortest.burst_time:
                    shortest = p
            order.append(shortest)
            remaining.remove(shortest)

        print("[SJF] Execution order (shortest burst first):")
        clock = 0
        for p in order:
            print(f"   t={clock:4d} -> {p.name} runs for {p.burst_time}ms")
            clock += p.burst_time
        print(f"[SJF] Total turnaround time: {clock}ms")
        return clock

    @staticmethod
    def round_robin_schedule(processes, quantum):
        print(f"[ROUND ROBIN] quantum={quantum}ms")
        queue = [[p, p.burst_time] for p in processes]
        clock = 0
        index = 0
        while queue:
            entry = queue[index]
            process, remaining = entry
            run_for = min(quantum, remaining)
            print(f"   t={clock:4d} -> {process.name} runs for {run_for}ms (remaining before: {remaining}ms)")
            clock += run_for
            remaining -= run_for
            if remaining <= 0:
                print(f"      {process.name} COMPLETE at t={clock}")
                queue.pop(index)
                if queue:
                    index %= len(queue)
            else:
                entry[1] = remaining
                index = (index + 1) % len(queue)
        print(f"[ROUND ROBIN] Total time: {clock}ms")
        return clock


# ==========================================================
# PROBLEM 3: TOURNAMENT PRACTICE-SLOT BOOKING (SORTED ARRAY + BINARY SEARCH)
# ==========================================================
class User:
    def __init__(self, user_id, name):
        self.user_id = user_id
        self.name = name

    def __str__(self):
        return self.name


class Resource:
    def __init__(self, resource_id, name):
        self.resource_id = resource_id
        self.name = name

    def __str__(self):
        return self.name


class Booking:
    def __init__(self, meeting_title, start_time, end_time, user=None, resource=None):
        self.title = meeting_title
        self.start = int(start_time)
        self.end = int(end_time)
        self.user = user
        self.resource = resource

    def __str__(self):
        who = f" by {self.user.name}" if self.user else ""
        return f"{self.title} ({self.start:04d} - {self.end:04d}){who}"


class ResourceScheduler:
    """Bookings for ONE resource, kept in a chronologically sorted array.
    A hand-written binary search locates the O(log n) insertion index.
    """

    def __init__(self):
        self.schedule = []

    def _binary_search_insertion_index(self, new_booking):
        low = 0
        high = len(self.schedule)
        print(f"   [BINARY SEARCH] Locating insertion index for start={new_booking.start:04d} "
              f"across {len(self.schedule)} existing booking(s).")

        while low < high:
            mid = (low + high) // 2
            mid_start = self.schedule[mid].start
            if mid_start < new_booking.start:
                print(f"      low={low}, high={high}, mid={mid} -> schedule[{mid}].start={mid_start:04d} "
                      f"< {new_booking.start:04d}, search RIGHT half")
                low = mid + 1
            else:
                print(f"      low={low}, high={high}, mid={mid} -> schedule[{mid}].start={mid_start:04d} "
                      f">= {new_booking.start:04d}, search LEFT half")
                high = mid

        print(f"   [BINARY SEARCH] Converged: insertion index = {low}")
        return low

    def request_booking(self, new_booking):
        print(f"[REQUEST] {new_booking}")
        insertion_index = self._binary_search_insertion_index(new_booking)

        if insertion_index > 0 and self.schedule[insertion_index - 1].end > new_booking.start:
            print(f"   [BOUNDARY CHECK] Left neighbour {self.schedule[insertion_index - 1]} "
                  f"ends after {new_booking.start:04d}.")
            print(f"   -> [REJECTED] Conflict with previous booking.")
            return False

        if insertion_index < len(self.schedule) and self.schedule[insertion_index].start < new_booking.end:
            print(f"   [BOUNDARY CHECK] Right neighbour {self.schedule[insertion_index]} "
                  f"starts before {new_booking.end:04d}.")
            print(f"   -> [REJECTED] Conflict with upcoming booking.")
            return False

        print(f"   [BOUNDARY CHECK] No overlap with neighbouring bookings.")
        self.schedule.insert(insertion_index, new_booking)
        print(f"   -> [SUCCESS] {new_booking.title} inserted at index {insertion_index} "
              f"(array elements shifted right to make space).")
        return True


class BookingSystem:
    """Manages many shared resources, each with its own sorted-array scheduler,
    so bookings on one resource never block another.
    """

    def __init__(self):
        self.resources = {}   # resource_id -> Resource
        self.schedulers = {}  # resource_id -> ResourceScheduler

    def add_resource(self, resource):
        self.resources[resource.resource_id] = resource
        self.schedulers[resource.resource_id] = ResourceScheduler()
        print(f"[RESOURCE] Registered shared resource '{resource}' (ID {resource.resource_id}).")

    def book(self, resource_id, user, title, start, end):
        if resource_id not in self.schedulers:
            print(f"[ERROR] Resource '{resource_id}' not found.")
            return False
        booking = Booking(title, start, end, user=user, resource=self.resources[resource_id])
        return self.schedulers[resource_id].request_booking(booking)


# ==========================================================
# PROBLEM 4: JD STOCK PARTS INVENTORY (HASH TABLE)
# ==========================================================
class InventoryItem:
    """part_no/category/price/notes are optional (real supplier data is often
    incomplete) so old-style calls that only pass item_id/name/current_stock/
    threshold still work unchanged.
    """

    def __init__(self, item_id, name, current_stock, threshold,
                 part_no="", category="", price=None, notes=""):
        self.item_id = item_id
        self.name = name
        self.current_stock = int(current_stock)
        self.threshold = int(threshold)
        self.part_no = part_no
        self.category = category
        self.price = price
        self.notes = notes

    def update_stock(self, quantity_change):
        self.current_stock += quantity_change

    def needs_reorder(self):
        return self.current_stock <= self.threshold

    def __str__(self):
        bits = [f"{self.name} (Stock: {self.current_stock}, Threshold: {self.threshold})"]
        if self.part_no:
            bits.append(f"[part no. {self.part_no}]")
        if self.category:
            bits.append(f"- {self.category}")
        if self.price is not None:
            bits.append(f"- Rs {self.price:g}")
        if self.notes:
            bits.append(f"({self.notes})")
        return " ".join(bits)


class InventoryManager:
    """JD Stock's real parts catalog, keyed by item_id in a Python dict (hash
    table), giving O(1) average-case direct access for stock adjustments.
    """

    def __init__(self):
        self.catalog = {}

    def register_item(self, new_item):
        self.catalog[new_item.item_id] = new_item
        print(f"[REGISTER] {new_item.name} stored under hash key '{new_item.item_id}'.")

    def adjust_stock(self, target_id, quantity_change):
        print(f"[LOOKUP] Direct O(1) hash access for key '{target_id}'.")
        if target_id in self.catalog:
            target_item = self.catalog[target_id]
            before = target_item.current_stock
            target_item.update_stock(quantity_change)
            print(f"[TRANSACTION] {target_item.name}: {before} -> {target_item.current_stock} "
                  f"(change: {quantity_change:+d})")
            if target_item.needs_reorder():
                print(f"   --> [ALERT] Critical stock level for {target_item.name}! "
                      f"({target_item.current_stock} <= threshold {target_item.threshold})")
        else:
            print(f"[ERROR] Item ID '{target_id}' not found.")

    def search_by_name(self, keyword):
        keyword_lower = keyword.lower()
        results = [item for item in self.catalog.values() if keyword_lower in item.name.lower()]
        print(f"[SEARCH] '{keyword}' matched {len(results)} item(s).")
        return results

    def sorted_by_stock(self):
        """Manual insertion sort, ascending by current_stock -- O(n^2), explicit
        for assessment purposes rather than relying on a built-in sort.
        """
        items = list(self.catalog.values())
        for i in range(1, len(items)):
            key_item = items[i]
            j = i - 1
            while j >= 0 and items[j].current_stock > key_item.current_stock:
                items[j + 1] = items[j]
                j -= 1
            items[j + 1] = key_item
        return items

    def auto_reorder(self, restock_amount=None):
        """Scan the catalog; any item at/below threshold is automatically restocked."""
        reordered = []
        for item in self.catalog.values():
            if item.needs_reorder():
                amount = restock_amount if restock_amount else item.threshold * 3
                print(f"[AUTO-REORDER] {item.name} at {item.current_stock} "
                      f"(<= threshold {item.threshold}). Placing reorder for {amount} units.")
                item.update_stock(amount)
                reordered.append(item)
        if not reordered:
            print("[AUTO-REORDER] No items require reordering.")
        return reordered

    def generate_report(self):
        width = 88
        print("=" * width)
        print(f"{'ID':<8}{'Name':<18}{'Part No.':<12}{'Category':<14}{'Stock':>7}{'Reorder@':>10}{'Status':>10}")
        print("-" * width)
        for item in self.sorted_by_stock():
            status = "REORDER" if item.needs_reorder() else "OK"
            print(f"{item.item_id:<8}{item.name:<18}{item.part_no:<12}{item.category:<14}"
                  f"{item.current_stock:>7}{item.threshold:>10}{status:>10}")
        print("=" * width)


# ==========================================================
# PROBLEM 5: JD STOCK PARTS RECOMMENDER (JACCARD SIMILARITY)
# ==========================================================
class Part:
    def __init__(self, part_id, name, category):
        self.part_id = part_id
        self.name = name
        self.category = category

    def __str__(self):
        return f"{self.name} [{self.category}]"


class Customer:
    def __init__(self, customer_id, name):
        self.customer_id = customer_id
        self.name = name

    def __str__(self):
        return self.name


class RecommendationEngine:
    """Collaborative filtering: customers who bought similar parts (by
    Jaccard similarity of their purchase sets) are used to recommend parts
    they haven't bought yet -- e.g. someone who bought Brake_Pads and
    Brake_Fluid the way most other Brake_Pads buyers did is a good candidate
    for whatever else that cohort tends to buy.
    """

    def __init__(self):
        self.parts = {}       # part_id -> Part
        self.customers = {}   # customer_id -> Customer
        self.purchases = {}   # customer_id -> set(part_id)

    def add_part(self, part):
        self.parts[part.part_id] = part
        print(f"[CATALOG] Added part '{part.part_id}' -> {part}")

    def add_customer(self, customer):
        self.customers[customer.customer_id] = customer
        self.purchases.setdefault(customer.customer_id, set())
        print(f"[CUSTOMER] Registered '{customer.customer_id}' -> {customer}")

    def record_purchase(self, customer_id, part_id):
        self.purchases.setdefault(customer_id, set()).add(part_id)
        buyer = self.customers.get(customer_id, customer_id)
        part = self.parts.get(part_id, part_id)
        print(f"[PURCHASE] {buyer} bought {part}")

    def jaccard_similarity(self, customer_a, customer_b):
        set_a = self.purchases.get(customer_a, set())
        set_b = self.purchases.get(customer_b, set())
        intersection = set_a & set_b
        union = set_a | set_b
        score = len(intersection) / len(union) if union else 0.0
        print(f"   [JACCARD] {customer_a} vs {customer_b}: |intersection|={len(intersection)}, "
              f"|union|={len(union)} -> similarity={score:.2f}")
        return score

    def find_similar_customers(self, target_customer, top_n=3):
        print(f"[SIMILARITY] Comparing '{target_customer}' against all other customers...")
        scores = []
        for other_id in self.purchases:
            if other_id == target_customer:
                continue
            scores.append((other_id, self.jaccard_similarity(target_customer, other_id)))
        scores.sort(key=lambda pair: pair[1], reverse=True)
        c = len(scores)
        print(f"[COMPLEXITY] Compared against {c} customer(s), each an O(p) set operation "
              f"-> O(c*p) for c customers and p parts/customer.")
        return scores[:top_n]

    def recommend(self, target_customer, top_n=3):
        already_owned = self.purchases.get(target_customer, set())
        similar_customers = self.find_similar_customers(target_customer)
        candidate_scores = {}
        for other_id, similarity in similar_customers:
            if similarity <= 0:
                continue
            for part_id in self.purchases.get(other_id, set()):
                if part_id not in already_owned:
                    candidate_scores[part_id] = candidate_scores.get(part_id, 0) + similarity

        ranked = sorted(candidate_scores.items(), key=lambda pair: pair[1], reverse=True)[:top_n]
        recommendations = [self.parts[part_id] for part_id, _ in ranked if part_id in self.parts]
        print(f"[RECOMMEND] Top picks for '{target_customer}': {[str(p) for p in recommendations]}")
        return recommendations


# ==========================================================
# CORRECTNESS TEST SUITE (unittest)
# ==========================================================
class TestDeliveryRouteOptimizer(unittest.TestCase):
    def test_route_visits_every_node_and_returns_to_start(self):
        opt = DeliveryRouteOptimizer()
        opt.add_location(Location("A", "Depot", 0, 0))
        opt.add_location(Location("B", "North", 0, 10))
        opt.add_location(Location("C", "East", 10, 0))
        opt.add_location(Location("D", "NE", 10, 10))
        route, distance = opt.nearest_neighbour_route("A")
        self.assertEqual(route[0], "A")
        self.assertEqual(route[-1], "A")
        self.assertEqual(set(route), {"A", "B", "C", "D"})
        self.assertGreater(distance, 0)

    def test_brute_force_matches_or_beats_nearest_neighbour(self):
        opt = DeliveryRouteOptimizer()
        opt.add_location(Location("A", "P0", 0, 0))
        opt.add_location(Location("B", "P1", 0, 5))
        opt.add_location(Location("C", "P2", 5, 5))
        opt.add_location(Location("D", "P3", 5, 0))
        _, nn_distance = opt.nearest_neighbour_route("A")
        _, optimal_distance = opt.brute_force_optimal("A")
        self.assertLessEqual(round(optimal_distance, 6), round(nn_distance, 6))

    def test_unknown_start_returns_none(self):
        opt = DeliveryRouteOptimizer()
        opt.add_location(Location("A", "P0", 0, 0))
        route, _ = opt.nearest_neighbour_route("Z")
        self.assertIsNone(route)


class TestTaskManager(unittest.TestCase):
    def test_process_allocated_when_resources_available(self):
        manager = TaskManager(total_cores=4, total_memory=1024)
        manager.submit_process("Render", priority=5, burst_time=100, cores_required=2, memory_required=256)
        manager.dispatch_all()
        self.assertEqual(len(manager.running), 1)
        self.assertEqual(manager.pool.available_cores, 2)

    def test_process_waits_when_resources_unavailable(self):
        manager = TaskManager(total_cores=1, total_memory=128)
        manager.submit_process("Big", priority=1, burst_time=50, cores_required=2, memory_required=64)
        manager.dispatch_all()
        self.assertEqual(len(manager.running), 0)
        self.assertEqual(len(manager.waiting), 1)

    def test_release_frees_resources_for_waiting_process(self):
        manager = TaskManager(total_cores=1, total_memory=128)
        manager.submit_process("First", priority=5, burst_time=50, cores_required=1, memory_required=64)
        manager.dispatch_all()
        manager.submit_process("Second", priority=1, burst_time=50, cores_required=1, memory_required=64)
        manager.dispatch_all()
        self.assertEqual(len(manager.waiting), 1)
        first_pid = manager.running[0].pid
        manager.release_process(first_pid)
        self.assertEqual(len(manager.waiting), 0)
        self.assertEqual(len(manager.running), 1)

    def test_higher_priority_dispatched_first(self):
        manager = TaskManager(total_cores=1, total_memory=64)
        manager.submit_process("Low", priority=1, burst_time=10, cores_required=1, memory_required=64)
        manager.submit_process("High", priority=9, burst_time=10, cores_required=1, memory_required=64)
        manager.dispatch_all()
        self.assertEqual(manager.running[0].name, "High")

    def test_sjf_orders_by_burst_time_ascending(self):
        procs = [Process(1, "Long", 1, 300, 1, 1), Process(2, "Short", 1, 50, 1, 1),
                 Process(3, "Medium", 1, 150, 1, 1)]
        total = TaskManager.sjf_schedule(procs)
        self.assertEqual(total, 500)


class TestBookingSystem(unittest.TestCase):
    def test_non_overlapping_booking_accepted(self):
        system = BookingSystem()
        system.add_resource(Resource("ROOM_A", "Meeting Room A"))
        user = User("u1", "Alice")
        self.assertTrue(system.book("ROOM_A", user, "Standup", 900, 1000))
        self.assertTrue(system.book("ROOM_A", user, "Retro", 1100, 1200))

    def test_overlapping_booking_rejected(self):
        system = BookingSystem()
        system.add_resource(Resource("ROOM_A", "Meeting Room A"))
        user = User("u1", "Alice")
        system.book("ROOM_A", user, "Standup", 900, 1000)
        self.assertFalse(system.book("ROOM_A", user, "Clash", 930, 1030))

    def test_bookings_on_different_resources_do_not_conflict(self):
        system = BookingSystem()
        system.add_resource(Resource("ROOM_A", "Room A"))
        system.add_resource(Resource("ROOM_B", "Room B"))
        user = User("u1", "Alice")
        self.assertTrue(system.book("ROOM_A", user, "Meeting", 900, 1000))
        self.assertTrue(system.book("ROOM_B", user, "Meeting", 900, 1000))

    def test_unknown_resource_rejected(self):
        system = BookingSystem()
        user = User("u1", "Alice")
        self.assertFalse(system.book("MISSING", user, "Meeting", 900, 1000))


class TestInventoryManager(unittest.TestCase):
    def test_register_and_lookup(self):
        inv = InventoryManager()
        inv.register_item(InventoryItem("SP_01", "Spark_Plugs", 50, 15))
        self.assertIn("SP_01", inv.catalog)
        self.assertEqual(inv.catalog["SP_01"].current_stock, 50)

    def test_stock_adjustment_updates_total(self):
        inv = InventoryManager()
        inv.register_item(InventoryItem("SP_01", "Spark_Plugs", 50, 15))
        inv.adjust_stock("SP_01", -5)
        self.assertEqual(inv.catalog["SP_01"].current_stock, 45)

    def test_reorder_alert_triggers_at_or_below_threshold(self):
        item = InventoryItem("BP_02", "Brake_Pads", 12, 10)
        self.assertFalse(item.needs_reorder())
        item.update_stock(-2)
        self.assertTrue(item.needs_reorder())

    def test_unknown_item_id_does_not_crash(self):
        inv = InventoryManager()
        inv.adjust_stock("MISSING", -1)

    def test_auto_reorder_brings_stock_above_threshold(self):
        inv = InventoryManager()
        inv.register_item(InventoryItem("BP_02", "Brake_Pads", 5, 10))
        inv.auto_reorder()
        self.assertFalse(inv.catalog["BP_02"].needs_reorder())

    def test_sorted_by_stock_is_ascending(self):
        inv = InventoryManager()
        inv.register_item(InventoryItem("A", "Item A", 50, 10))
        inv.register_item(InventoryItem("B", "Item B", 5, 10))
        inv.register_item(InventoryItem("C", "Item C", 20, 10))
        stocks = [item.current_stock for item in inv.sorted_by_stock()]
        self.assertEqual(stocks, sorted(stocks))

    def test_search_by_name_is_case_insensitive_substring(self):
        inv = InventoryManager()
        inv.register_item(InventoryItem("SP_01", "Spark_Plugs", 50, 15))
        results = inv.search_by_name("spark")
        self.assertEqual(len(results), 1)


class TestRecommendationEngine(unittest.TestCase):
    def _build_engine(self):
        engine = RecommendationEngine()
        engine.add_customer(Customer("C1", "Ramesh Motors"))
        engine.add_customer(Customer("C2", "Kalanki Garage"))
        engine.add_part(Part("SP_01", "Spark_Plugs", "Engine"))
        engine.add_part(Part("BP_02", "Brake_Pads", "Brakes"))
        engine.add_part(Part("BF_03", "Brake_Fluid", "Brakes"))
        engine.record_purchase("C1", "SP_01")
        engine.record_purchase("C1", "BP_02")
        engine.record_purchase("C2", "SP_01")
        engine.record_purchase("C2", "BP_02")
        engine.record_purchase("C2", "BF_03")
        return engine

    def test_jaccard_similarity_known_value(self):
        engine = self._build_engine()
        score = engine.jaccard_similarity("C1", "C2")
        self.assertAlmostEqual(score, 2 / 3)

    def test_recommend_excludes_already_owned_parts(self):
        engine = self._build_engine()
        recommendations = engine.recommend("C1")
        part_ids = [p.part_id for p in recommendations]
        self.assertNotIn("SP_01", part_ids)
        self.assertNotIn("BP_02", part_ids)
        self.assertIn("BF_03", part_ids)
def run_scripted_demo():
    print("========================================")
    print("PROBLEM 1: JD STOCK DELIVERY ROUTE OPTIMIZER")
    print("========================================")
    optimizer = DeliveryRouteOptimizer()
    optimizer.add_location(Location("WH", "Kathmandu_Warehouse", 0, 0))
    optimizer.add_location(Location("PKR", "Pokhara_Branch", 0, 8))
    optimizer.add_location(Location("BTW", "Butwal_Customer", 6, 0))
    optimizer.add_location(Location("BRT", "Biratnagar_Customer", 6, 8))
    route, distance = optimizer.nearest_neighbour_route("WH")
    print(f"Nearest-Neighbour route: {' -> '.join(route)}  Total distance: {distance:.2f}")
    best_route, best_distance = optimizer.brute_force_optimal("WH")
    print(f"Brute-force optimal:     {' -> '.join(best_route)}  Total distance: {best_distance:.2f}")

    print("\n========================================")
    print("PROBLEM 2: ZULU SERVER JOB SCHEDULING")
    print("========================================")
    manager = TaskManager(total_cores=4, total_memory=1024)
    manager.submit_process("Match_Recap_Generation", priority=70, burst_time=200, cores_required=2, memory_required=512)
    manager.submit_process("Tournament_Reminder_Sweep", priority=30, burst_time=100, cores_required=1, memory_required=256)
    manager.submit_process("Room_Code_Broadcast", priority=99, burst_time=20, cores_required=2, memory_required=400)
    manager.dispatch_all()
    print("\nComparing FCFS vs SJF vs Round Robin on the same three jobs...")
    demo_processes = [Process(1, "Match_Recap_Generation", 70, 200, 2, 512),
                       Process(2, "Tournament_Reminder_Sweep", 30, 100, 1, 256),
                       Process(3, "Room_Code_Broadcast", 99, 20, 2, 400)]
    TaskManager.fcfs_schedule(demo_processes)
    TaskManager.sjf_schedule(demo_processes)
    TaskManager.round_robin_schedule(demo_processes, quantum=50)

    print("\n========================================")
    print("PROBLEM 3: TOURNAMENT PRACTICE-SLOT BOOKING")
    print("========================================")
    booking_system = BookingSystem()
    booking_system.add_resource(Resource("PRACTICE_1", "Practice_Server_1"))
    alice = User("u1", "Squad_Alpha")
    booking_system.book("PRACTICE_1", alice, "Scrim_Warmup", 900, 1000)
    booking_system.book("PRACTICE_1", alice, "Bracket_Rehearsal", 1100, 1200)
    booking_system.book("PRACTICE_1", alice, "Late_Scrim", 1015, 1045)     # Valid
    booking_system.book("PRACTICE_1", alice, "Overlapping_Request", 1030, 1130)  # Conflict

    print("\n========================================")
    print("PROBLEM 4: JD STOCK PARTS INVENTORY")
    print("========================================")
    inv_manager = InventoryManager()
    inv_manager.register_item(InventoryItem("SP_01", "Spark_Plugs", 50, 15, part_no="NGK-BPR6ES", category="Engine", price=180))
    inv_manager.register_item(InventoryItem("BP_02", "Brake_Pads", 12, 10, part_no="BP-4402", category="Brakes", price=1450))
    inv_manager.adjust_stock("SP_01", -5)
    inv_manager.adjust_stock("BP_02", -4)
    inv_manager.auto_reorder()
    inv_manager.generate_report()

    print("\n========================================")
    print("PROBLEM 5: JD STOCK PARTS RECOMMENDER")
    print("========================================")
    engine = RecommendationEngine()
    engine.add_customer(Customer("C1", "Ramesh_Motors"))
    engine.add_customer(Customer("C2", "Kalanki_Garage"))
    engine.add_part(Part("SP_01", "Spark_Plugs", "Engine"))
    engine.add_part(Part("BP_02", "Brake_Pads", "Brakes"))
    engine.add_part(Part("BF_03", "Brake_Fluid", "Brakes"))
    engine.record_purchase("C1", "SP_01")
    engine.record_purchase("C1", "BP_02")
    engine.record_purchase("C2", "SP_01")
    engine.record_purchase("C2", "BP_02")
    engine.record_purchase("C2", "BF_03")
    engine.recommend("C1")

    print("\n========================================\nALL DEMOS COMPLETE.")

def prompt_int(label):
    while True:
        raw = input(label).strip()
        try:
            return int(raw)
        except ValueError:
            print(f"   Please enter a whole number (got '{raw}').")


def route_optimizer_menu(optimizer):
    while True:
        print("\n--- PROBLEM 1: JD STOCK DELIVERY ROUTES ---")
        if optimizer.locations:
            for loc in optimizer.locations.values():
                print(f"   {loc.location_id}: {loc}")
        else:
            print("   (no locations yet)")
        print("1) Add location  2) Nearest-Neighbour route  3) Brute-force optimum  4) Back")
        choice = input("Choose: ").strip()
        if choice == "1":
            loc_id = input("Location ID: ").strip()
            name = input("Location name: ").strip()
            x = prompt_int("X coordinate: ")
            y = prompt_int("Y coordinate: ")
            optimizer.add_location(Location(loc_id, name, x, y))
        elif choice == "2":
            if len(optimizer.locations) < 2:
                print("   Need at least 2 locations.")
                continue
            start_id = input("Start location ID: ").strip()
            route, distance = optimizer.nearest_neighbour_route(start_id)
            if route:
                print(f"Route: {' -> '.join(route)}  Total distance: {distance:.2f}")
        elif choice == "3":
            start_id = input("Start location ID: ").strip()
            route, distance = optimizer.brute_force_optimal(start_id)
            if route:
                print(f"Optimal route: {' -> '.join(route)}  Total distance: {distance:.2f}")
        elif choice == "4":
            break
        else:
            print("   Invalid option.")


def resource_allocation_menu(manager):
    while True:
        print("\n--- PROBLEM 2: ZULU SERVER JOB SCHEDULING ---")
        print(f"   Pool: {manager.pool}")
        print(f"   Running: {[f'{p.name}(PID{p.pid})' for p in manager.running]}")
        print(f"   Waiting: {[p.name for p in manager.waiting]}")
        print("1) Submit process  2) Dispatch queued processes  3) Release a running process")
        print("4) Compare FCFS vs SJF vs Round Robin  5) Back")
        choice = input("Choose: ").strip()
        if choice == "1":
            name = input("Process name: ").strip()
            priority = prompt_int("Priority (higher = more urgent): ")
            burst = prompt_int("Burst time (ms): ")
            cores = prompt_int("Cores required: ")
            memory = prompt_int("Memory required (MB): ")
            manager.submit_process(name, priority, burst, cores, memory)
        elif choice == "2":
            manager.dispatch_all()
        elif choice == "3":
            pid = prompt_int("PID to release: ")
            manager.release_process(pid)
        elif choice == "4":
            n = prompt_int("How many demo processes to simulate: ")
            demo = [Process(i, f"Demo{i}", (i * 7) % 10 + 1, (i * 37) % 200 + 20, 1, 64)
                    for i in range(1, n + 1)]
            TaskManager.fcfs_schedule(demo)
            TaskManager.sjf_schedule(demo)
            TaskManager.round_robin_schedule(demo, quantum=50)
        elif choice == "5":
            break
        else:
            print("   Invalid option.")

def booking_system_menu(booking_system):
    while True:
        print("\n--- PROBLEM 3: TOURNAMENT PRACTICE-SLOT BOOKING ---")
        if booking_system.resources:
            for rid, resource in booking_system.resources.items():
                sched = booking_system.schedulers[rid]
                print(f"   {rid} ({resource}): [{', '.join(str(b) for b in sched.schedule)}]")
        else:
            print("   (no shared resources registered yet)")
        print("1) Register shared resource  2) Request booking  3) Back")
        choice = input("Choose: ").strip()
        if choice == "1":
            rid = input("Resource ID: ").strip()
            name = input("Resource name: ").strip()
            booking_system.add_resource(Resource(rid, name))
        elif choice == "2":
            rid = input("Resource ID to book: ").strip()
            user_name = input("Your name: ").strip()
            title = input("Booking title: ").strip()
            start = prompt_int("Start time (e.g. 900): ")
            end = prompt_int("End time (e.g. 1000): ")
            if end <= start:
                print("   End time must be after start time.")
                continue
            booking_system.book(rid, User(user_name, user_name), title, start, end)
        elif choice == "3":
            break
        else:
            print("   Invalid option.")


def inventory_menu(manager):
    while True:
        print("\n--- PROBLEM 4: JD STOCK PARTS INVENTORY ---")
        if manager.catalog:
            for item in manager.catalog.values():
                print(f"   {item.item_id}: {item}")
        else:
            print("   (catalog is empty)")
        print("1) Register item  2) Adjust stock  3) Search by name  4) Show sorted-by-stock")
        print("5) Auto-reorder low-stock items  6) Generate report  7) Back")
        choice = input("Choose: ").strip()
        if choice == "1":
            item_id = input("Item ID (e.g. SP_01): ").strip()
            name = input("Item name: ").strip()
            stock = prompt_int("Current stock: ")
            threshold = prompt_int("Reorder threshold: ")
            part_no = input("Part no. (optional): ").strip()
            category = input("Category (optional): ").strip()
            price_raw = input("Price in Rs (optional): ").strip()
            price = float(price_raw) if price_raw else None
            notes = input("Notes (optional): ").strip()
            manager.register_item(InventoryItem(item_id, name, stock, threshold,
                                                  part_no=part_no, category=category,
                                                  price=price, notes=notes))
        elif choice == "2":
            item_id = input("Item ID to adjust: ").strip()
            change = prompt_int("Quantity change (e.g. -5, +10): ")
            manager.adjust_stock(item_id, change)
        elif choice == "3":
            keyword = input("Search keyword: ").strip()
            for item in manager.search_by_name(keyword):
                print(f"   {item.item_id}: {item}")
        elif choice == "4":
            for item in manager.sorted_by_stock():
                print(f"   {item.item_id}: {item}")
        elif choice == "5":
            manager.auto_reorder()
        elif choice == "6":
            manager.generate_report()
        elif choice == "7":
            break
        else:
            print("   Invalid option.")


def recommendation_menu(engine):
    while True:
        print("\n--- PROBLEM 5: JD STOCK PARTS RECOMMENDER ---")
        print(f"   Customers: {[str(c) for c in engine.customers.values()]}")
        print(f"   Parts: {[str(p) for p in engine.parts.values()]}")
        print("1) Add part  2) Add customer  3) Record purchase  4) Recommend parts for a customer  5) Back")
        choice = input("Choose: ").strip()
        if choice == "1":
            part_id = input("Part ID: ").strip()
            name = input("Part name: ").strip()
            category = input("Category: ").strip()
            engine.add_part(Part(part_id, name, category))
        elif choice == "2":
            customer_id = input("Customer ID: ").strip()
            name = input("Customer name: ").strip()
            engine.add_customer(Customer(customer_id, name))
        elif choice == "3":
            customer_id = input("Customer ID: ").strip()
            part_id = input("Part ID: ").strip()
            engine.record_purchase(customer_id, part_id)
        elif choice == "4":
            customer_id = input("Customer ID to recommend for: ").strip()
            engine.recommend(customer_id)
        elif choice == "5":
            break
        else:
            print("   Invalid option.")


# JD Stock's real starter catalog -- Spark_Plugs and Brake_Pads were the only two items
# actually on record; the rest is a realistic-sized catalog for a Nepal-based bike/car
# parts shop built out around them (same categories, same part-number/pricing shape),
# so the inventory and recommender demos look like an actual 5-years-running business
# instead of two rows. Fields: item_id, name, part_no, category, price (Rs), current
# stock, reorder threshold, notes. A handful (AF_02, BD_10, ALT_14, VBELT_45) sit at or
# below their threshold on purpose, so the reorder-alert path has something to show.
JD_STOCK_CATALOG = [
    ("SP_01", "Spark_Plugs", "NGK-BPR6ES", "Engine", 180, 42, 20, "Fits Bajaj Pulsar 150/180, Hero Splendor"),
    ("AF_02", "Air_Filter", "K&N-BM-9910", "Engine", 950, 8, 10, ""),
    ("OF_03", "Oil_Filter", "Hiflofiltro-HF204", "Engine", 320, 25, 12, ""),
    ("FF_04", "Fuel_Filter", "Hero-OEM-16010", "Engine", 260, 14, 8, ""),
    ("EO_05", "Engine_Oil_20W50_1L", "Castrol-Power1", "Engine", 550, 60, 25, "Best seller, 4-stroke bikes"),
    ("TC_06", "Timing_Chain_Kit", "Bajaj-OEM-2203", "Engine", 1450, 6, 5, ""),
    ("HG_07", "Head_Gasket", "Bajaj-OEM-2540", "Engine", 680, 9, 6, ""),
    ("VBELT_45", "Fan_Belt", "Bajaj-OEM-2601", "Engine", 260, 4, 6, ""),
    ("PST_46", "Piston_Ring_Set", "Bajaj-OEM-2205", "Engine", 890, 5, 4, ""),
    ("BP_02", "Brake_Pads", "BP-4402", "Brakes", 1450, 12, 10, ""),
    ("BS_08", "Brake_Shoes", "Hero-OEM-3301", "Brakes", 480, 18, 10, ""),
    ("BF_03", "Brake_Fluid_DOT4", "Motul-RBF600", "Brakes", 620, 22, 10, ""),
    ("BD_10", "Brake_Disc_Front", "Bajaj-OEM-2710", "Brakes", 1850, 4, 5, ""),
    ("BC_11", "Brake_Cable", "Hero-OEM-4102", "Brakes", 220, 30, 12, ""),
    ("BCL_12", "Brake_Caliper", "TVS-OEM-5590", "Brakes", 2400, 3, 3, ""),
    ("BAT_13", "Battery_12V_5Ah", "Exide-Xplore", "Electrical", 2800, 7, 6, ""),
    ("ALT_14", "Alternator", "Bosch-0123XX", "Electrical", 6500, 2, 3, ""),
    ("STM_15", "Starter_Motor", "Bosch-0001XX", "Electrical", 3200, 4, 3, ""),
    ("HB_16", "Headlight_Bulb_H4", "Philips-H4", "Electrical", 350, 40, 15, ""),
    ("TL_17", "Tail_Light_Bulb", "Generic-BA15", "Electrical", 90, 55, 20, ""),
    ("FUS_18", "Fuse_Kit", "Generic-Assorted", "Electrical", 150, 33, 15, ""),
    ("HRN_19", "Horn", "Roots-Falcon", "Electrical", 480, 11, 8, ""),
    ("IND_47", "Indicator_Bulb", "Generic", "Electrical", 60, 46, 20, ""),
    ("SHF_20", "Shock_Absorber_Front", "Bajaj-OEM-3105", "Suspension", 1650, 6, 4, ""),
    ("SHR_21", "Shock_Absorber_Rear", "Bajaj-OEM-3106", "Suspension", 1800, 5, 4, ""),
    ("CS_22", "Coil_Spring", "Generic-CS100", "Suspension", 900, 10, 5, ""),
    ("TRE_23", "Tie_Rod_End", "TVS-OEM-6620", "Suspension", 750, 9, 6, ""),
    ("BJ_24", "Ball_Joint", "TVS-OEM-6621", "Suspension", 680, 8, 6, ""),
    ("SRB_25", "Steering_Rack_Boot", "Generic-SRB200", "Suspension", 320, 12, 8, ""),
    ("TY_26", "Tyre_100_90_17", "MRF-Nylogrip", "Tyres", 3200, 16, 8, "Most common bike size in Nepal"),
    ("TB_27", "Tube_100_90_17", "CEAT", "Tyres", 450, 28, 12, ""),
    ("TY2_28", "Tyre_155_65_R13", "MRF-ZVTS", "Tyres", 5200, 5, 4, "Fits Maruti Alto/WagonR"),
    ("WBW_29", "Wheel_Balancing_Weight", "Generic", "Tyres", 40, 120, 40, ""),
    ("GO_30", "Gear_Oil_80W90_1L", "Castrol", "Fluids", 480, 20, 10, ""),
    ("CL_31", "Chain_Lubricant", "Motul-C2", "Fluids", 650, 15, 8, ""),
    ("GRS_32", "Multipurpose_Grease", "Servo", "Fluids", 220, 24, 10, ""),
    ("COOL_44", "Coolant_1L", "Castrol", "Fluids", 380, 17, 8, ""),
    ("RAD_33", "Radiator", "Bajaj-OEM-4410", "Cooling", 3800, 3, 3, ""),
    ("RC_34", "Radiator_Cap", "Generic", "Cooling", 180, 19, 8, ""),
    ("CF_35", "Cooling_Fan", "Bosch", "Cooling", 1200, 5, 4, ""),
    ("CP_36", "Clutch_Plate_Set", "Hero-OEM-5501", "Transmission", 1350, 7, 5, ""),
    ("CCB_37", "Clutch_Cable", "Hero-OEM-5502", "Transmission", 210, 26, 12, ""),
    ("CVT_38", "CVT_Belt", "TVS-Jupiter-OEM", "Transmission", 1450, 6, 5, "Scooter drivebelt"),
    ("SM_39", "Side_Mirror", "Generic-Universal", "Body", 280, 34, 15, ""),
    ("WB_40", "Wiper_Blade", "Bosch-Aerotwin", "Body", 650, 13, 8, ""),
    ("MG_41", "Mudguard_Front", "Bajaj-OEM-6001", "Body", 420, 10, 6, ""),
    ("SPK_42", "Chain_Sprocket_Kit", "RK-Chains", "Body", 2100, 8, 5, ""),
    ("SEAT_43", "Seat_Cover_Universal", "Generic", "Body", 380, 21, 10, ""),
]

# Which customer garages have bought which parts, built from realistic co-purchase
# patterns (bike-only garages cluster together, car garages cluster together) so
# Problem 5's Jaccard similarity has something meaningful to compare.
JD_STOCK_CUSTOMERS = {
    "Ramesh_Motors": ["SP_01", "AF_02", "OF_03", "EO_05", "BS_08", "BC_11", "CCB_37"],
    "Kalanki_Garage": ["SP_01", "BP_02", "BF_03", "HB_16", "CL_31", "TY_26", "TB_27"],
    "Balaju_Auto_Works": ["TY2_28", "RAD_33", "ALT_14", "WB_40", "BF_03", "BD_10", "STM_15"],
    "New_Baneshwor_Bike_Center": ["SP_01", "OF_03", "EO_05", "CVT_38", "SPK_42", "TY_26"],
    "Pokhara_Motor_House": ["SP_01", "BP_02", "TY2_28", "RC_34", "GO_30", "CS_22"],
    "Butwal_Spare_Parts": ["HB_16", "TL_17", "FUS_18", "IND_47", "GRS_32"],
}


def load_sample_data(route_optimizer, task_manager, booking_system, inv_manager, rec_engine):
    """Populates all five systems with example data in one step, so the menus
    aren't empty on first use. Same data as the scripted demo, but written
    into the live objects you'll actually browse/extend from the menu.
    """
    route_optimizer.add_location(Location("WH", "Kathmandu_Warehouse", 0, 0))
    route_optimizer.add_location(Location("PKR", "Pokhara_Branch", 0, 8))
    route_optimizer.add_location(Location("BTW", "Butwal_Customer", 6, 0))
    route_optimizer.add_location(Location("BRT", "Biratnagar_Customer", 6, 8))

    task_manager.submit_process("Match_Recap_Generation", priority=70, burst_time=200, cores_required=2, memory_required=512)
    task_manager.submit_process("Tournament_Reminder_Sweep", priority=30, burst_time=100, cores_required=1, memory_required=256)
    task_manager.submit_process("Room_Code_Broadcast", priority=99, burst_time=20, cores_required=2, memory_required=400)
    task_manager.dispatch_all()

    booking_system.add_resource(Resource("PRACTICE_1", "Practice_Server_1"))
    alice = User("u1", "Squad_Alpha")
    booking_system.book("PRACTICE_1", alice, "Scrim_Warmup", 900, 1000)
    booking_system.book("PRACTICE_1", alice, "Bracket_Rehearsal", 1100, 1200)

    for item_id, name, part_no, category, price, stock, threshold, notes in JD_STOCK_CATALOG:
        inv_manager.register_item(InventoryItem(item_id, name, stock, threshold,
                                                  part_no=part_no, category=category,
                                                  price=price, notes=notes))
        rec_engine.add_part(Part(item_id, name, category))

    for i, (customer_name, part_ids) in enumerate(JD_STOCK_CUSTOMERS.items(), start=1):
        customer_id = f"C{i}"
        rec_engine.add_customer(Customer(customer_id, customer_name))
        for part_id in part_ids:
            rec_engine.record_purchase(customer_id, part_id)

    print("\n[SAMPLE DATA] Loaded into all five systems -- open menus 1-5 to explore them.")


def save_data_to_excel(filename, route_optimizer, task_manager, booking_system, inv_manager, rec_engine):
    """Serialises the live state of all five systems into one Excel workbook
    (one sheet per system) so it survives between runs of the program.
    """
    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.active
    ws.title = "Locations"
    ws.append(["location_id", "name", "x", "y"])
    for loc in route_optimizer.locations.values():
        ws.append([loc.location_id, loc.name, loc.x, loc.y])

    ws = wb.create_sheet("ResourcePool")
    ws.append(["total_cores", "total_memory", "available_cores", "available_memory", "next_pid"])
    pool = task_manager.pool
    ws.append([pool.total_cores, pool.total_memory, pool.available_cores, pool.available_memory,
               task_manager._next_pid])

    ws = wb.create_sheet("Processes")
    ws.append(["state", "pid", "name", "priority", "burst_time", "cores_required", "memory_required"])
    for p in task_manager.ready_heap.heap:
        ws.append(["queued", p.pid, p.name, p.priority, p.burst_time, p.cores_required, p.memory_required])
    for p in task_manager.waiting:
        ws.append(["waiting", p.pid, p.name, p.priority, p.burst_time, p.cores_required, p.memory_required])
    for p in task_manager.running:
        ws.append(["running", p.pid, p.name, p.priority, p.burst_time, p.cores_required, p.memory_required])

    ws = wb.create_sheet("Resources")
    ws.append(["resource_id", "name"])
    for r in booking_system.resources.values():
        ws.append([r.resource_id, r.name])

    ws = wb.create_sheet("Bookings")
    ws.append(["resource_id", "title", "start", "end", "user_id", "user_name"])
    for resource_id, scheduler in booking_system.schedulers.items():
        for b in scheduler.schedule:
            ws.append([resource_id, b.title, b.start, b.end,
                       b.user.user_id if b.user else "", b.user.name if b.user else ""])

    ws = wb.create_sheet("Inventory")
    ws.append(["item_id", "name", "current_stock", "threshold", "part_no", "category", "price", "notes"])
    for item in inv_manager.catalog.values():
        ws.append([item.item_id, item.name, item.current_stock, item.threshold,
                   item.part_no, item.category, item.price, item.notes])

    ws = wb.create_sheet("Customers")
    ws.append(["customer_id", "name"])
    for c in rec_engine.customers.values():
        ws.append([c.customer_id, c.name])

    ws = wb.create_sheet("Parts")
    ws.append(["part_id", "name", "category"])
    for p in rec_engine.parts.values():
        ws.append([p.part_id, p.name, p.category])

    ws = wb.create_sheet("Purchases")
    ws.append(["customer_id", "part_id"])
    for customer_id, part_ids in rec_engine.purchases.items():
        for part_id in part_ids:
            ws.append([customer_id, part_id])

    wb.save(filename)
    print(f"[SAVE] Data saved to '{filename}'.")


def load_data_from_excel(filename, route_optimizer, task_manager, booking_system, inv_manager, rec_engine):
    """Restores all five systems from a workbook previously written by
    save_data_to_excel(). Returns True if a file was found and loaded.
    """
    if not os.path.exists(filename):
        return False

    from openpyxl import load_workbook
    wb = load_workbook(filename)

    def rows(sheet_name):
        if sheet_name not in wb.sheetnames:
            return
        for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            yield row

    for location_id, name, x, y in rows("Locations"):
        route_optimizer.add_location(Location(location_id, name, x, y))

    pool_row = next(rows("ResourcePool"), None)
    if pool_row:
        total_cores, total_memory, available_cores, available_memory, next_pid = pool_row
        pool = task_manager.pool
        pool.total_cores, pool.total_memory = total_cores, total_memory
        pool.available_cores, pool.available_memory = available_cores, available_memory
        task_manager._next_pid = next_pid

    for state, pid, name, priority, burst_time, cores_required, memory_required in rows("Processes"):
        process = Process(pid, name, priority, burst_time, cores_required, memory_required)
        if state == "queued":
            task_manager.ready_heap.insert(process)
        elif state == "waiting":
            task_manager.waiting.append(process)
        elif state == "running":
            task_manager.running.append(process)

    for resource_id, name in rows("Resources"):
        booking_system.add_resource(Resource(resource_id, name))

    for resource_id, title, start, end, user_id, user_name in rows("Bookings"):
        if resource_id not in booking_system.schedulers:
            continue
        user = User(user_id, user_name) if user_id else None
        booking = Booking(title, start, end, user=user, resource=booking_system.resources.get(resource_id))
        booking_system.schedulers[resource_id].schedule.append(booking)
    for scheduler in booking_system.schedulers.values():
        scheduler.schedule.sort(key=lambda b: b.start)

    for item_id, name, current_stock, threshold, part_no, category, price, notes in rows("Inventory"):
        inv_manager.register_item(InventoryItem(
            item_id, name, current_stock, threshold,
            part_no=part_no or "", category=category or "", price=price, notes=notes or ""))

    for customer_id, name in rows("Customers"):
        rec_engine.add_customer(Customer(customer_id, name))

    for part_id, name, category in rows("Parts"):
        rec_engine.add_part(Part(part_id, name, category))

    for customer_id, part_id in rows("Purchases"):
        rec_engine.record_purchase(customer_id, part_id)

    print(f"[LOAD] Data loaded from '{filename}'.")
    return True


def main_menu():
    route_optimizer = DeliveryRouteOptimizer()
    task_manager = TaskManager(total_cores=8, total_memory=16384)
    booking_system = BookingSystem()
    inv_manager = InventoryManager()
    rec_engine = RecommendationEngine()

    print("========================================")
    print("JD Systems Toolkit -- delivery routes, server jobs, practice bookings,")
    print("JD Stock inventory and JD Stock recommendations, in one place.")
    print("========================================")
    print("Five systems. Pick 1-5 to add your own data and try each one live,")
    print("or 8 to pre-load examples instantly.")

    try:
        restored = load_data_from_excel(DATA_FILE, route_optimizer, task_manager, booking_system,
                                         inv_manager, rec_engine)
    except Exception as exc:
        print(f"[STARTUP] Could not read '{DATA_FILE}' ({exc}). Starting fresh with empty systems.")
        restored = False

    if restored:
        print(f"[STARTUP] Restored your previous session from '{DATA_FILE}'.")
    else:
        print(f"[STARTUP] No saved data found. Starting fresh -- will save to '{DATA_FILE}'.")

    while True:
        print("\n========================================")
        print("MAIN MENU")
        print("========================================")
        print("1) Problem 1: JD Stock Delivery Routes (Graph + Nearest Neighbour TSP)")
        print("2) Problem 2: Zulu Server Job Scheduling (Max-Heap + FCFS/SJF/Round Robin)")
        print("3) Problem 3: Tournament Practice-Slot Booking (Sorted Array + Binary Search)")
        print("4) Problem 4: JD Stock Parts Inventory (Hash Table)")
        print("5) Problem 5: JD Stock Parts Recommender (Jaccard Similarity)")
        print("6) Run scripted demo (fixed trace, for screenshots)")
        print("7) Run correctness test suite")
        print("8) Load sample data into all systems (quick start)")
        print("0) Exit")
        choice = input("Choose: ").strip()

        if choice == "1":
            route_optimizer_menu(route_optimizer)
            save_data_to_excel(DATA_FILE, route_optimizer, task_manager, booking_system, inv_manager, rec_engine)
        elif choice == "2":
            resource_allocation_menu(task_manager)
            save_data_to_excel(DATA_FILE, route_optimizer, task_manager, booking_system, inv_manager, rec_engine)
        elif choice == "3":
            booking_system_menu(booking_system)
            save_data_to_excel(DATA_FILE, route_optimizer, task_manager, booking_system, inv_manager, rec_engine)
        elif choice == "4":
            inventory_menu(inv_manager)
            save_data_to_excel(DATA_FILE, route_optimizer, task_manager, booking_system, inv_manager, rec_engine)
        elif choice == "5":
            recommendation_menu(rec_engine)
            save_data_to_excel(DATA_FILE, route_optimizer, task_manager, booking_system, inv_manager, rec_engine)
        elif choice == "6":
            run_scripted_demo()
        elif choice == "7":
            suite = unittest.TestLoader().loadTestsFromModule(__import__("__main__"))
            unittest.TextTestRunner(verbosity=2).run(suite)
        elif choice == "8":
            load_sample_data(route_optimizer, task_manager, booking_system, inv_manager, rec_engine)
            save_data_to_excel(DATA_FILE, route_optimizer, task_manager, booking_system, inv_manager, rec_engine)
        elif choice == "0":
            save_data_to_excel(DATA_FILE, route_optimizer, task_manager, booking_system, inv_manager, rec_engine)
            print("Goodbye.")
            break
        else:
            print("   Invalid option.")


# ==========================================================
# MASTER EXECUTION BLOCK
# ==========================================================
if __name__ == "__main__":
    main_menu()
