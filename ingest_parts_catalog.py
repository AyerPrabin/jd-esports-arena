"""
INGEST PARTS CATALOG  —  one-off/manual admin script, NOT part of the running server
=====================================================================================
Run this by hand (`python ingest_parts_catalog.py`) whenever a supplier parts catalog
(.xlsx or .pdf) is dropped into jdstock_catalog/. It's a best-effort heuristic extractor,
not a universal parser — real-world catalogs vary wildly in layout, so it prints a
per-file summary of what it found so you can sanity-check (or skip/hand-fix) a file that
didn't extract cleanly, rather than silently feeding ZULU garbage.

Output is cached to zulu_parts_catalog.json (git-ignored) so zulu_parts.py doesn't have
to re-parse PDFs/Excels on every server start — parsing only happens when you re-run this.
"""
import glob
import json
import os
import re

CATALOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "jdstock_catalog")
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "zulu_parts_catalog.json")

# Header keywords we try to match each column/row-label against, in priority order.
_HEADER_HINTS = {
    "part_no": ["part no", "part number", "partno", "sku", "code", "item no", "item code"],
    "name": ["part name", "description", "item name", "item", "name", "product"],
    "category": ["category", "type", "group", "section"],
    "price": ["price", "rate", "mrp", "cost", "amount", "rs", "npr"],
    "notes": ["notes", "remarks", "compatible", "model", "fitment", "note"],
}


def _classify_header(cell_text):
    """Match a header cell's text against known field hints. Returns the field name or None."""
    t = str(cell_text or "").strip().lower()
    if not t:
        return None
    for field, hints in _HEADER_HINTS.items():
        if any(h in t for h in hints):
            return field
    return None


def _parse_price(v):
    if v is None:
        return None
    s = str(v)
    m = re.search(r"[\d,]+\.?\d*", s)
    if not m:
        return None
    try:
        return float(m.group(0).replace(",", ""))
    except ValueError:
        return None


def _rows_from_table(table, source_file):
    """table: list of rows (each a list of cell values). First row assumed to be headers.
    Returns list of row-dicts; skips a table if it can't find at least a name-like column,
    since a table with no identifiable name column isn't safely usable for lookups."""
    if not table or len(table) < 2:
        return []
    header_row = table[0]
    col_field = {i: _classify_header(c) for i, c in enumerate(header_row)}
    if "name" not in col_field.values():
        return []
    rows = []
    for raw in table[1:]:
        if not raw or all(c in (None, "") for c in raw):
            continue
        row = {"name": "", "part_no": "", "price": None, "category": "", "notes": "",
               "source_file": source_file}
        for i, cell in enumerate(raw):
            field = col_field.get(i)
            if not field or cell in (None, ""):
                continue
            if field == "price":
                row["price"] = _parse_price(cell)
            else:
                existing = row.get(field, "")
                row[field] = (existing + " " + str(cell)).strip() if existing else str(cell).strip()
        if row["name"]:
            rows.append(row)
    return rows


def _parse_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    table = [list(row) for row in sheet.iter_rows(values_only=True)]
    wb.close()
    return _rows_from_table(table, os.path.basename(path))


def _parse_pdf(path):
    import pdfplumber
    rows = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                rows.extend(_rows_from_table(table, os.path.basename(path)))
    return rows


def ingest():
    os.makedirs(CATALOG_DIR, exist_ok=True)
    files = sorted(glob.glob(os.path.join(CATALOG_DIR, "*.xlsx")) +
                    glob.glob(os.path.join(CATALOG_DIR, "*.pdf")))
    if not files:
        print(f"No .xlsx/.pdf files found in {CATALOG_DIR} — nothing to ingest.")
        return

    all_rows = []
    for path in files:
        name = os.path.basename(path)
        try:
            if path.lower().endswith(".xlsx"):
                rows = _parse_xlsx(path)
            else:
                rows = _parse_pdf(path)
        except Exception as e:
            print(f"  [FAIL] {name}: {e}")
            continue
        if rows:
            print(f"  [OK]   {name}: {len(rows)} parts extracted "
                  f"(e.g. \"{rows[0]['name']}\"{' @ ' + str(rows[0]['price']) if rows[0]['price'] else ''})")
        else:
            print(f"  [SKIP] {name}: no recognizable parts table found — check its layout "
                  f"or add a header row with columns like Name/Part No/Price")
        all_rows.extend(rows)

    tmp = OUT_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(all_rows, f, ensure_ascii=False, indent=2)
    os.replace(tmp, OUT_PATH)
    print(f"\nWrote {len(all_rows)} total parts from {len(files)} file(s) to {OUT_PATH}")


if __name__ == "__main__":
    ingest()
